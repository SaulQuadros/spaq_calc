import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from io import BytesIO

TEMPLATE_PATH = Path("11_Dimens_SPAQ_AQ_v1.xlsx")

st.set_page_config(page_title="SPAQ/AQ -> Streamlit", layout="wide")
st.title("Dimensionamento SPAQ/AQ")

st.markdown("""
Este aplicativo reproduz os quadros principais da planilha **11_Dimens_SPAQ_AQ_v1.xlsx**.
- As células _azuis_ do Excel são **entradas editáveis**.
- A coluna **Peso total** é **bloqueada** e apresenta o cálculo automático.
- Linhas de somas e totais não aparecem dentro das tabelas; os resultados são exibidos abaixo.
""")

def read_initial_tables(path: Path):
    if path.exists():
        try:
            wb = load_workbook(path, data_only=False)
            ws = wb.active
            t1_rows, t2_rows = [], []
            r = 3
            while r <= ws.max_row and ws.cell(row=r, column=1).value not in [None, ""]:
                t1_rows.append({
                    "Aparelho": ws.cell(row=r, column=1).value,
                    "Vazão (L/min)": ws.cell(row=r, column=2).value,
                    "Pressão (m.c.a)": ws.cell(row=r, column=3).value,
                    "Quantidade": ws.cell(row=r, column=4).value,
                    "Peso": ws.cell(row=r, column=5).value
                })
                r += 1
            r = 10
            while r <= ws.max_row and ws.cell(row=r, column=1).value not in [None, ""]:
                t2_rows.append({
                    "Aparelho": ws.cell(row=r, column=1).value,
                    "Vazão (L/min)": ws.cell(row=r, column=2).value,
                    "Pressão (m.c.a)": ws.cell(row=r, column=3).value,
                    "Quantidade": ws.cell(row=r, column=4).value,
                    "Peso": ws.cell(row=r, column=5).value
                })
                r += 1
            params = {k: ws[k].value for k in ["B3","B18","B19","B20","B26","B31","B32","B33","B35","C41","C42","C43"]}
            return pd.DataFrame(t1_rows), pd.DataFrame(t2_rows), params
        except Exception as e:
            st.warning(f"Erro ao ler o arquivo Excel: {e}. Usando dados padrão.")

    # Dados padrão se o arquivo não existir
    t1 = pd.DataFrame([
        {"Aparelho": "Chuveiro", "Vazão (L/min)": 12, "Pressão (m.c.a)": 4, "Quantidade": 5, "Peso": 0.4},
        {"Aparelho": "Lavatório", "Vazão (L/min)": 8, "Pressão (m.c.a)": 4, "Quantidade": 5, "Peso": 0.2},
        {"Aparelho": "Pia de cozinha", "Vazão (L/min)": 8, "Pressão (m.c.a)": 4, "Quantidade": 2, "Peso": 0.2}
    ])
    t2 = pd.DataFrame([
        {"Aparelho": "Tanque de lavar roupas", "Vazão (L/min)": 15, "Pressão (m.c.a)": 2, "Quantidade": 1, "Peso": 0.7},
        {"Aparelho": "Máquina de lavar roupas", "Vazão (L/min)": 15, "Pressão (m.c.a)": 2, "Quantidade": 1, "Peso": 0.7},
        {"Aparelho": "Vaso sanitário", "Vazão (L/min)": 8, "Pressão (m.c.a)": 2, "Quantidade": 1, "Peso": 0.5}
    ])
    params = {"B3": 12, "B18": 45, "B19": 20, "B20": 40, "B26": 0.8, "B31": 5, "B32": 6, "B33": 2, "B35": 5,
              "C41": 2, "C42": 21, "C43": 483}
    return t1, t2, params

t1_init, t2_init, params_init = read_initial_tables(TEMPLATE_PATH)

st.sidebar.header("Parâmetros gerais")
B3 = st.sidebar.number_input("B3 (chuveiros por ramal)", value=float(params_init.get("B3", 12)))
B18 = st.sidebar.number_input("B18 (Pressão disponível m.c.a)", value=float(params_init.get("B18", 45)))
B19 = st.sidebar.number_input("B19 (Pressão de consumo m.c.a)", value=float(params_init.get("B19", 20)))
B20 = st.sidebar.number_input("B20 (Perda adicional m.c.a)", value=float(params_init.get("B20", 40)))
B26 = st.sidebar.number_input("B26 (Eficiência ou divisor)", value=float(params_init.get("B26", 0.8)))
B31 = st.sidebar.number_input("B31", value=float(params_init.get("B31", 5)))
B32 = st.sidebar.number_input("B32", value=float(params_init.get("B32", 6)))
B33 = st.sidebar.number_input("B33", value=float(params_init.get("B33", 2)))
B35 = st.sidebar.number_input("B35", value=float(params_init.get("B35", 5)))

# ---- TABELA 1: Aparelhos com AF e AQ ----
st.header("Aparelhos com AF e AQ")
t1 = t1_init.copy()
for col in ["Vazão (L/min)", "Pressão (m.c.a)", "Quantidade", "Peso"]:
    t1[col] = pd.to_numeric(t1[col], errors="coerce").fillna(0)
t1["Peso total"] = (t1["Quantidade"] * t1["Peso"]).round(4)

t1_edit = st.data_editor(
    t1,
    num_rows="dynamic",
    disabled=["Peso total"],  # bloqueia edição da coluna calculada
    key="t1_editor"
)

F6 = (t1_edit["Quantidade"] * t1_edit["Peso"]).sum()
F7 = round(60 * (0.3 * max(F6, 0) ** 0.5), 1)

st.markdown("**Resultados — Aparelhos com AF e AQ**")
col1, col2 = st.columns(2)
col1.metric("Soma dos Pesos (F6)", f"{F6:.3f}")
col2.metric("Vazão total (F7)", f"{F7:.1f} L/h")

# ---- TABELA 2: Aparelhos só AF ----
st.header("Aparelhos só AF")
t2 = t2_init.copy()
for col in ["Vazão (L/min)", "Pressão (m.c.a)", "Quantidade", "Peso"]:
    t2[col] = pd.to_numeric(t2[col], errors="coerce").fillna(0)
t2["Peso total"] = (t2["Quantidade"] * t2["Peso"]).round(4)

t2_edit = st.data_editor(
    t2,
    num_rows="dynamic",
    disabled=["Peso total"],
    key="t2_editor"
)

F13 = (t2_edit["Quantidade"] * t2_edit["Peso"]).sum()
F14 = round(60 * (0.3 * max(F13, 0) ** 0.5), 1)

st.markdown("**Resultados — Aparelhos só AF**")
col3, col4 = st.columns(2)
col3.metric("Soma dos Pesos (F13)", f"{F13:.3f}")
col4.metric("Vazão total (F14)", f"{F14:.1f} L/h")

# ---- CÁLCULOS COMBINADOS ----
F15 = round(60 * (0.3 * max(F6 + F13, 0) ** 0.5) * 0.06, 1)
B21 = 1 - (B18 - B20) / (B18 - B19) if (B18 - B19) != 0 else 0
B22 = F7 * B21
B25 = B22 * (B18 - B19)
B27 = B25 / B26 if B26 != 0 else 0
B30 = max(t1_edit["Pressão (m.c.a)"].max(), 0)
B34 = B30 + B31 + B32 + B33
B36 = B35 - B34
C39 = F15
C40 = 0 if B36 > 0 else -1 * B36
C41 = st.sidebar.number_input("Aquecedor - Quantidade (C41)", value=float(params_init.get("C41", 2)))
C42 = st.sidebar.number_input("Aquecedor - Vazão (L/min) (C42)", value=float(params_init.get("C42", 21)))
C43 = st.sidebar.number_input("Aquecedor - Potência (kcal/min) (C43)", value=float(params_init.get("C43", 483)))
C44 = (C43 * C41 / (B18 - B19) / B3) if (B18 - B19) != 0 and B3 != 0 else 0

st.markdown("### Indicadores combinados")
cols_comb = st.columns(3)
with cols_comb[0]:
    st.metric("Vazão combinada (F15)", f"{F15:.2f}")
    st.metric("B22 (Q ajustado)", f"{B22:.2f}")
with cols_comb[1]:
    st.metric("B27 (Resultado)", f"{B27:.2f}")
    st.metric("C39 (Pressurizador Q)", f"{C39:.2f}")
with cols_comb[2]:
    st.metric("C40 (Altura manométrica)", f"{C40:.2f}")
    st.metric("Qt. Chuveiros (C44)", f"{C44:.2f}")

# --- Botão de download do Excel ---
output = BytesIO()
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    t1_edit.to_excel(writer, index=False, sheet_name="Aparelhos_AF_AQ")
    t2_edit.to_excel(writer, index=False, sheet_name="Aparelhos_AF")
    resumo = pd.DataFrame({
        "Indicador": ["F6", "F7", "F13", "F14", "F15", "B36", "C44"],
        "Valor": [F6, F7, F13, F14, F15, B36, C44]
    })
    resumo.to_excel(writer, index=False, sheet_name="Resumo")
output.seek(0)

st.download_button(
    label="📥 Baixar resultados em Excel",
    data=output.getvalue(),
    file_name="Resultados_SPAQ.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.success("App pronto para execução no Streamlit Cloud.")