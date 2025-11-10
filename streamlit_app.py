
import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE_PATH = Path("11_Dimens_SPAQ_AQ_v1.xlsx")

st.set_page_config(page_title="SPAQ/AQ -> Streamlit", layout="wide")

st.title("Conversão: Planilha → App (Streamlit)")
st.markdown("""
Este aplicativo reproduz os quadros principais da planilha **11_Dimens_SPAQ_AQ_v1.xlsx**.
- As células _azuis_ do Excel foram tratadas como **entradas do usuário**.
- As células com fórmulas foram reimplementadas em Python (pandas / numpy) e exibidas como **resultados calculados**.
- Você pode adicionar/remover linhas nas tabelas "Aparelhos com AF e AQ" e "Aparelhos só AF".
""")

# --- Helper to read initial tables from the template ---
def read_initial_tables(path):
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    # Table 1 header at row 2, data rows start at 3 until a blank in column A
    t1_rows = []
    r = 3
    while r <= ws.max_row and ws.cell(row=r, column=1).value is not None:
        label = ws.cell(row=r, column=1).value
        vazao = ws.cell(row=r, column=2).value
        press = ws.cell(row=r, column=3).value
        qtd = ws.cell(row=r, column=4).value
        peso = ws.cell(row=r, column=5).value
        t1_rows.append({"Aparelho": label, "Vazão (L/min)": vazao, "Pressão (m.c.a)": press, "Quantidade": qtd, "Peso": peso})
        r += 1
    # Table 2 header at row 9, data rows start at 10
    t2_rows = []
    r = 10
    while r <= ws.max_row and ws.cell(row=r, column=1).value is not None:
        label = ws.cell(row=r, column=1).value
        vazao = ws.cell(row=r, column=2).value
        press = ws.cell(row=r, column=3).value
        qtd = ws.cell(row=r, column=4).value
        peso = ws.cell(row=r, column=5).value
        t2_rows.append({"Aparelho": label, "Vazão (L/min)": vazao, "Pressão (m.c.a)": press, "Quantidade": qtd, "Peso": peso})
        r += 1
    # Read other numeric parameters referenced by formulas
    params = {}
    params["B3"] = ws["B3"].value
    params["B18"] = ws["B18"].value
    params["B19"] = ws["B19"].value
    params["B20"] = ws["B20"].value
    params["B26"] = ws["B26"].value
    params["B31"] = ws["B31"].value
    params["B32"] = ws["B32"].value
    params["B33"] = ws["B33"].value
    params["B35"] = ws["B35"].value
    # Aquecedor related
    params["C41"] = ws["C41"].value
    params["C42"] = ws["C42"].value
    params["C43"] = ws["C43"].value
    return pd.DataFrame(t1_rows), pd.DataFrame(t2_rows), params

t1_init, t2_init, params_init = read_initial_tables(TEMPLATE_PATH)

st.sidebar.header("Parâmetros gerais (edição rápida)")
B3 = st.sidebar.number_input("B3 (divisor - ex: chuveiros por ramal)", value=float(params_init.get("B3",12)))
B18 = st.sidebar.number_input("B18 (Pressão disponível m.c.a)", value=float(params_init.get("B18",45)))
B19 = st.sidebar.number_input("B19 (Pressão de consumo m.c.a)", value=float(params_init.get("B19",20)))
B20 = st.sidebar.number_input("B20 (Perda adicional m.c.a)", value=float(params_init.get("B20",40)))
B26 = st.sidebar.number_input("B26 (Eficiência ou divisor)", value=float(params_init.get("B26",0.8)))
B31 = st.sidebar.number_input("B31 (valor fixo 1)", value=float(params_init.get("B31",5)))
B32 = st.sidebar.number_input("B32 (valor fixo 2)", value=float(params_init.get("B32",6)))
B33 = st.sidebar.number_input("B33 (valor fixo 3)", value=float(params_init.get("B33",2)))
B35 = st.sidebar.number_input("B35 (pressão necessária disponível)", value=float(params_init.get("B35",5)))

st.header("Aparelhos com AF e AQ")
st.markdown("Edite os valores das células 'azuis' (Vazão, Pressão, Quantidade, Peso). Adicione ou remova linhas conforme necessário.")
t1 = st.experimental_data_editor(t1_init, num_rows="dynamic", key="t1_editor")

st.header("Aparelhos só AF")
st.markdown("Edite os valores das células 'azuis' (Vazão, Pressão, Quantidade, Peso). Adicione ou remova linhas conforme necessário.")
t2 = st.experimental_data_editor(t2_init, num_rows="dynamic", key="t2_editor")

# Ensure numeric types
for df in (t1, t2):
    for col in ["Vazão (L/min)", "Pressão (m.c.a)", "Quantidade", "Peso"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# --- Calculations (replicando as fórmulas encontradas na planilha) ---
# Peso total por linha = Quantidade * Peso
t1["Peso total"] = t1["Quantidade"] * t1["Peso"]
t2["Peso total"] = t2["Quantidade"] * t2["Peso"]

# Somas
F6 = t1["Peso total"].sum()  # sum F3:F5 originally
F13 = t2["Peso total"].sum()  # sum F10:F12 originally

# Vazões calculadas (segundo fórmulas da planilha)
F7 = round(60 * (0.3 * (F6 if F6>0 else 0) ** 0.5), 1)
F14 = round(60 * (0.3 * (F13 if F13>0 else 0) ** 0.5), 1)
F15 = round(60 * (0.3 * ((F6 + F13) if (F6+F13)>0 else 0) ** 0.5) * 0.06, 1)

# Pressure & derived
B21 = 1 - (B18 - B20) / (B18 - B19) if (B18 - B19) != 0 else 0
B22 = F7 * B21
B25 = B22 * (B18 - B19)
B27 = B25 / B26 if B26 != 0 else 0

B30 = max(t1["Pressão (m.c.a)"].max() if len(t1)>0 else 0, 0)
B34 = B30 + B31 + B32 + B33
B36 = B35 - B34

C39 = F15
C40 = 0 if B36 > 0 else -1 * B36

# Aquecedor final calc
C41 = st.sidebar.number_input("Aquecedor - Quantidade (C41)", value=float(params_init.get("C41",2)))
C42 = st.sidebar.number_input("Aquecedor - Vazão (L/min) (C42)", value=float(params_init.get("C42",21)))
C43 = st.sidebar.number_input("Aquecedor - Potência (kcal/min) (C43)", value=float(params_init.get("C43",483)))

C44 = (C43 * C41 / (B18 - B19) / B3) if (B18 - B19) != 0 and B3 != 0 else 0

# --- Display results ---
st.subheader("Resultados — Tabela 1 (Aparelhos com AF e AQ)")
st.dataframe(t1.set_index("Aparelho"))

st.subheader("Resultados — Tabela 2 (Aparelhos só AF)")
st.dataframe(t2.set_index("Aparelho"))

st.subheader("Cálculos resumidos")
cols = st.columns(3)
with cols[0]:
    st.metric("F6 (Peso total T1)", f"{F6:.2f}")
    st.metric("F13 (Peso total T2)", f"{F13:.2f}")
    st.metric("F7 (Vazão T1 L/h)", f"{F7:.1f}")
with cols[1]:
    st.metric("F14 (Vazão T2 L/h)", f"{F14:.1f}")
    st.metric("F15 (Vazão combinada m³/h?)", f"{F15:.1f}")
    st.metric("B27 (resultado)", f"{B27:.2f}")
with cols[2]:
    st.metric("B36 (Excedente/falta pressão)", f"{B36:.2f}")
    st.metric("C39 (Pressurizador Q)", f"{C39:.2f}")
    st.metric("C40 (Altura man. H)", f"{C40:.2f}")

st.markdown("---")
st.write("Cálculo do número de chuveiros recomendado (C44):")
st.write(f"Qt. Chuveiros ≈ {C44:.2f} (com base em C43={C43}, C41={C41}, B18={B18}, B19={B19}, B3={B3})")

st.info("Observação: Reimplementei as fórmulas detectadas na planilha. Se houver fórmulas adicionais que queira reproduzir, posso adicioná-las.")

st.markdown("### Exportar resultados")
if st.button("Salvar resultados em novo Excel (gera arquivo local)"):
    # build a new workbook from template and write inputs + computed columns
    import openpyxl, io
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    # write back table1 starting at row 3
    start_row = 3
    for idx, row in t1.reset_index(drop=True).iterrows():
        r = start_row + idx
        ws.cell(row=r, column=1, value=row["Aparelho"])
        ws.cell(row=r, column=2, value=row["Vazão (L/min)"])
        ws.cell(row=r, column=3, value=row["Pressão (m.c.a)"])
        ws.cell(row=r, column=4, value=row["Quantidade"])
        ws.cell(row=r, column=5, value=row["Peso"])
        ws.cell(row=r, column=6, value=float(row["Peso total"]))
    # table2 starting at row 10
    start_row = 10
    for idx, row in t2.reset_index(drop=True).iterrows():
        r = start_row + idx
        ws.cell(row=r, column=1, value=row["Aparelho"])
        ws.cell(row=r, column=2, value=row["Vazão (L/min)"])
        ws.cell(row=r, column=3, value=row["Pressão (m.c.a)"])
        ws.cell(row=r, column=4, value=row["Quantidade"])
        ws.cell(row=r, column=5, value=row["Peso"])
        ws.cell(row=r, column=6, value=float(row["Peso total"]))
    # write summary cells
    ws["F6"] = F6
    ws["F7"] = F7
    ws["F13"] = F13
    ws["F14"] = F14
    ws["F15"] = F15
    ws["B21"] = B21
    ws["B22"] = B22
    ws["B25"] = B25
    ws["B27"] = B27
    ws["B30"] = B30
    ws["B34"] = B34
    ws["B36"] = B36
    ws["C39"] = C39
    ws["C40"] = C40
    ws["C44"] = C44
    out_path = Path("results_spaq_output.xlsx")
    wb.save(out_path)
    st.success(f"Arquivo salvo: {out_path} (baixe do diretório do app)")
